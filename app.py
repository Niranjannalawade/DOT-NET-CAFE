import os
import re
import json
import glob
import mimetypes
import shutil
from datetime import datetime
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, send_from_directory, send_file, jsonify)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cybercafe_secret_2025')

DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///cybercafe.db')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
UPLOAD_BASE = os.path.join(BASE_DIR, 'static', 'uploads')
EXCEL_DIR   = os.path.join(BASE_DIR, 'static', 'exports')

# ── Scanner watch-folder config file ──────────────────────────────────────────
# Stores the path where the scanner software saves files.
# Example:  C:\Users\Admin\Documents\Scans
# Saved as a plain text file so it persists across restarts.
SCANNER_CONFIG_FILE = os.path.join(BASE_DIR, 'scanner_path.txt')

app.config['UPLOAD_BASE']        = UPLOAD_BASE
app.config['EXCEL_DIR']          = EXCEL_DIR
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024   # 16 MB

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'docx', 'doc', 'txt', 'gif'}

ICON_MAP = {
    'pdf':  ('fa-file-pdf',   'icon-pdf'),
    'png':  ('fa-file-image', 'icon-img'),
    'jpg':  ('fa-file-image', 'icon-img'),
    'jpeg': ('fa-file-image', 'icon-img'),
    'gif':  ('fa-file-image', 'icon-img'),
    'docx': ('fa-file-word',  'icon-word'),
    'doc':  ('fa-file-word',  'icon-word'),
    'txt':  ('fa-file-lines', 'icon-txt'),
}

db = SQLAlchemy(app)


# ─── Models ───────────────────────────────────────────────────────────────────

class Customer(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100), nullable=False)
    phone      = db.Column(db.String(20),  nullable=False)
    email      = db.Column(db.String(120))
    place      = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    documents  = db.relationship('Document', backref='customer', lazy=True, cascade='all, delete-orphan')
    activities = db.relationship('Activity', backref='customer', lazy=True, cascade='all, delete-orphan')
    bills      = db.relationship('Bill',     backref='customer', lazy=True, cascade='all, delete-orphan')


class Document(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    customer_id   = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    filename      = db.Column(db.String(255), nullable=False)
    original_name = db.Column(db.String(255))
    file_size     = db.Column(db.Integer, default=0)
    upload_date   = db.Column(db.DateTime, default=datetime.utcnow)

    @property
    def ext(self):
        return self.filename.rsplit('.', 1)[-1].lower() if '.' in self.filename else ''

    @property
    def icon(self):
        return ICON_MAP.get(self.ext, ('fa-file', 'icon-default'))

    @property
    def size_str(self):
        s = self.file_size or 0
        if s < 1024:        return f'{s} B'
        if s < 1048576:     return f'{s/1024:.1f} KB'
        return f'{s/1048576:.1f} MB'

    @property
    def is_image(self):
        return self.ext in {'png', 'jpg', 'jpeg', 'gif'}


class Activity(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    description = db.Column(db.String(500), nullable=False)
    date        = db.Column(db.DateTime, default=datetime.utcnow)


class Bill(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    invoice_number = db.Column(db.String(20), unique=True, nullable=False)
    customer_id    = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    items          = db.Column(db.Text, nullable=False)
    subtotal       = db.Column(db.Float, default=0)
    discount       = db.Column(db.Float, default=0)
    total          = db.Column(db.Float, default=0)
    date           = db.Column(db.DateTime, default=datetime.utcnow)


# ─── Scanner path helpers ──────────────────────────────────────────────────────

def get_scanner_path():
    """
    Read the saved scanner watch-folder path from scanner_path.txt.
    Returns empty string if file not found or empty.
    """
    try:
        with open(SCANNER_CONFIG_FILE, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except FileNotFoundError:
        return ''


def set_scanner_path(path):
    """
    Save the scanner watch-folder path to scanner_path.txt.
    """
    with open(SCANNER_CONFIG_FILE, 'w', encoding='utf-8') as f:
        f.write(path.strip())


def get_latest_scanned_file(folder_path):
    """
    Look inside folder_path for the MOST RECENTLY MODIFIED file
    with an allowed extension.

    Returns the full file path of the newest file, or None if
    the folder is empty or no allowed files are found.

    This is how the "Scan from folder" feature works:
      1. Scanner saves a file to the watch folder
      2. This function finds the newest file in that folder
      3. Flask copies it to the customer's upload folder
    """
    if not folder_path or not os.path.isdir(folder_path):
        return None

    # Collect all allowed files in the folder
    found = []
    for ext in ALLOWED_EXTENSIONS:
        # Match both lower and upper case extensions
        found += glob.glob(os.path.join(folder_path, f'*.{ext}'))
        found += glob.glob(os.path.join(folder_path, f'*.{ext.upper()}'))

    if not found:
        return None

    # Return the file with the most recent modification time
    return max(found, key=os.path.getmtime)


# ─── General helpers ───────────────────────────────────────────────────────────

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def sanitize_name(name):
    """
    "Bhavesh Oswal" → "Bhavesh_Oswal"
    "Riya (Shah)"   → "Riya_Shah"
    """
    name = name.strip()
    name = re.sub(r'\s+', '_', name)
    name = re.sub(r'[^\w]', '', name)
    name = re.sub(r'_+', '_', name)
    name = name.strip('_')
    return name if name else 'customer'


def customer_upload_dir(customer_id):
    """
    Return (and auto-create) static/uploads/<CustomerName>/
    """
    c = Customer.query.get(customer_id)
    folder_name = sanitize_name(c.name) if c else f'customer_{customer_id}'
    folder = os.path.join(app.config['UPLOAD_BASE'], folder_name)
    os.makedirs(folder, exist_ok=True)
    return folder


def _save_file_object(file_obj, original_name, customer_id):
    """
    Save a werkzeug FileStorage object to the customer folder.
    Returns the new Document (not yet committed).
    Raises ValueError on validation failure.
    """
    if not file_obj or not original_name:
        raise ValueError('No file received.')
    if not allowed_file(original_name):
        ext = original_name.rsplit('.', 1)[-1].upper() if '.' in original_name else '?'
        raise ValueError(f'".{ext}" files are not allowed. Use PDF, JPG, PNG, DOCX or TXT.')

    timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name   = secure_filename(original_name)
    stored_name = f'{timestamp}_{safe_name}'
    folder      = customer_upload_dir(customer_id)
    full_path   = os.path.join(folder, stored_name)

    file_obj.save(full_path)

    doc = Document(
        customer_id   = customer_id,
        filename      = stored_name,
        original_name = original_name,
        file_size     = os.path.getsize(full_path),
    )
    db.session.add(doc)
    db.session.add(Activity(
        customer_id = customer_id,
        description = f'Document uploaded: {original_name} ({doc.size_str})',
    ))
    return doc


def _save_disk_file(src_path, customer_id):
    """
    Copy an EXISTING file on the server's disk (e.g. from the scanner
    watch folder) into the customer's upload folder.

    Returns the new Document (not yet committed).
    Raises ValueError on validation failure.
    """
    original_name = os.path.basename(src_path)

    if not allowed_file(original_name):
        ext = original_name.rsplit('.', 1)[-1].upper() if '.' in original_name else '?'
        raise ValueError(f'".{ext}" files are not allowed.')

    if not os.path.isfile(src_path):
        raise ValueError(f'File not found on scanner path: {src_path}')

    timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name   = secure_filename(original_name)
    stored_name = f'{timestamp}_{safe_name}'
    folder      = customer_upload_dir(customer_id)
    full_path   = os.path.join(folder, stored_name)

    shutil.copy2(src_path, full_path)   # copy2 preserves metadata

    doc = Document(
        customer_id   = customer_id,
        filename      = stored_name,
        original_name = original_name,
        file_size     = os.path.getsize(full_path),
    )
    db.session.add(doc)
    db.session.add(Activity(
        customer_id = customer_id,
        description = f'Scanned document imported: {original_name} ({doc.size_str})',
    ))
    return doc


def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_next_invoice_number():
    last = Bill.query.order_by(Bill.id.desc()).first()
    num  = (int(last.invoice_number.split('-')[1]) + 1) if last else 1
    return f'INV-{num:05d}'


def rebuild_excel():
    customers = Customer.query.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = ['ID', 'Name', 'Phone', 'Email', 'Place', 'Registered On']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.fill      = PatternFill(start_color='16213E', end_color='16213E', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    for row, c in enumerate(customers, 2):
        data = [c.id, c.name, c.phone, c.email or '', c.place or '',
                c.created_at.strftime('%Y-%m-%d') if c.created_at else '']
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val).alignment = Alignment(horizontal='center')
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    os.makedirs(app.config['EXCEL_DIR'], exist_ok=True)
    path = os.path.join(app.config['EXCEL_DIR'], 'customers.xlsx')
    wb.save(path)
    return path


# ─── Auth ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard') if session.get('logged_in') else url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form.get('username') == 'admin' and request.form.get('password') == 'admin123':
            session['logged_in'] = True
            return redirect(url_for('dashboard'))
        flash('Invalid username or password.', 'error')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ─── Dashboard ────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html',
        total_customers   = Customer.query.count(),
        total_bills       = Bill.query.count(),
        total_revenue     = db.session.query(db.func.sum(Bill.total)).scalar() or 0,
        recent_bills      = Bill.query.order_by(Bill.date.desc()).limit(5).all(),
        recent_customers  = Customer.query.order_by(Customer.created_at.desc()).limit(5).all(),
        recent_activities = Activity.query.order_by(Activity.date.desc()).limit(8).all(),
    )


# ─── Customer CRUD ────────────────────────────────────────────────────────────

@app.route('/customers')
@login_required
def customers():
    q     = request.args.get('q', '')
    custs = (Customer.query
             .filter((Customer.name.ilike(f'%{q}%')) | (Customer.phone.ilike(f'%{q}%')))
             .all()) if q else Customer.query.order_by(Customer.created_at.desc()).all()
    return render_template('customers.html', customers=custs, q=q)


@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
def add_customer():
    if request.method == 'POST':
        c = Customer(
            name  = request.form['name'],
            phone = request.form['phone'],
            email = request.form.get('email', ''),
            place = request.form.get('place', ''),
        )
        db.session.add(c)
        db.session.commit()
        customer_upload_dir(c.id)
        db.session.add(Activity(customer_id=c.id, description='Customer registered'))
        db.session.commit()
        rebuild_excel()
        flash('Customer added successfully!', 'success')
        return redirect(url_for('customers'))
    return render_template('customer_form.html', customer=None)


@app.route('/customers/edit/<int:cid>', methods=['GET', 'POST'])
@login_required
def edit_customer(cid):
    c = Customer.query.get_or_404(cid)
    if request.method == 'POST':
        old_folder = customer_upload_dir(cid)
        c.name  = request.form['name']
        c.phone = request.form['phone']
        c.email = request.form.get('email', '')
        c.place = request.form.get('place', '')
        db.session.commit()
        new_folder = os.path.join(app.config['UPLOAD_BASE'], sanitize_name(c.name))
        if old_folder != new_folder and os.path.isdir(old_folder):
            os.makedirs(new_folder, exist_ok=True)
            for fname in os.listdir(old_folder):
                shutil.move(os.path.join(old_folder, fname),
                            os.path.join(new_folder, fname))
            try:
                os.rmdir(old_folder)
            except OSError:
                pass
        db.session.add(Activity(customer_id=c.id, description='Customer details updated'))
        db.session.commit()
        rebuild_excel()
        flash('Customer updated successfully!', 'success')
        return redirect(url_for('customers'))
    return render_template('customer_form.html', customer=c)


@app.route('/customers/delete/<int:cid>', methods=['POST'])
@login_required
def delete_customer(cid):
    c = Customer.query.get_or_404(cid)
    folder = customer_upload_dir(cid)
    if os.path.isdir(folder):
        shutil.rmtree(folder)
    db.session.delete(c)
    db.session.commit()
    rebuild_excel()
    flash('Customer and all data deleted.', 'success')
    return redirect(url_for('customers'))


# ─── Customer Profile ─────────────────────────────────────────────────────────

@app.route('/customers/profile/<int:cid>')
@login_required
def customer_profile(cid):
    c          = Customer.query.get_or_404(cid)
    documents  = Document.query.filter_by(customer_id=cid).order_by(Document.upload_date.desc()).all()
    activities = Activity.query.filter_by(customer_id=cid).order_by(Activity.date.desc()).all()
    bills      = Bill.query.filter_by(customer_id=cid).order_by(Bill.date.desc()).all()

    # Get scanner path for display in the UI
    scanner_path = get_scanner_path()

    # If scanner path is configured, peek at the latest file for preview
    latest_scan  = None
    if scanner_path:
        lf = get_latest_scanned_file(scanner_path)
        if lf:
            latest_scan = {
                'name':     os.path.basename(lf),
                'modified': datetime.fromtimestamp(os.path.getmtime(lf)).strftime('%d %b %Y, %H:%M:%S'),
                'size':     _human_size(os.path.getsize(lf)),
            }

    return render_template('profile.html',
        customer     = c,
        documents    = documents,
        activities   = activities,
        bills        = bills,
        total_spent  = sum(b.total for b in bills),
        folder_name  = sanitize_name(c.name),
        scanner_path = scanner_path,
        latest_scan  = latest_scan,
    )


def _human_size(b):
    if b < 1024:      return f'{b} B'
    if b < 1048576:   return f'{b/1024:.1f} KB'
    return f'{b/1048576:.1f} MB'


# ─── Scanner Path Config ───────────────────────────────────────────────────────
# GET  /scanner/config   → returns JSON with current path
# POST /scanner/config   → saves new path, returns JSON

@app.route('/scanner/config', methods=['GET', 'POST'])
@login_required
def scanner_config():
    """
    GET  → return current scanner path as JSON
    POST → save new scanner path, return JSON confirmation

    Called by JavaScript on the settings / profile page.
    """
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        new_path = data.get('path', '').strip()

        if not new_path:
            return jsonify(ok=False, message='Path cannot be empty.'), 400

        # Validate the path exists on the server machine
        if not os.path.isdir(new_path):
            return jsonify(
                ok      = False,
                message = f'Folder not found: "{new_path}". '
                          f'Make sure the scanner software saves files to this exact path.'
            ), 400

        set_scanner_path(new_path)
        return jsonify(ok=True, message=f'Scanner folder set to: {new_path}', path=new_path)

    # GET
    path = get_scanner_path()
    return jsonify(ok=True, path=path, configured=bool(path))


# ─── Scan from Folder (the key new route) ─────────────────────────────────────
# POST /scan_from_folder/<customer_id>
#
# Flow:
#   1. Read scanner watch-folder path from scanner_path.txt
#   2. Find the newest file in that folder
#   3. Copy it to static/uploads/<CustomerName>/
#   4. Save DB record + activity log
#   5. Return JSON → JS reloads the page

@app.route('/scan_from_folder/<int:customer_id>', methods=['POST'])
@login_required
def scan_from_folder(customer_id):
    """
    Import the latest scanned file from the configured scanner folder
    directly into the customer's document folder.

    No file is uploaded from the browser — the file is already on the
    server disk (saved there by the scanner software).
    """
    c = Customer.query.get(customer_id)
    if not c:
        return jsonify(ok=False, message='Customer not found.'), 404

    # 1. Get scanner path
    scanner_path = get_scanner_path()
    if not scanner_path:
        return jsonify(
            ok      = False,
            message = 'Scanner folder not configured. '
                      'Please set the scanner path in Settings first.',
            action  = 'open_settings'    # JS uses this to open settings panel
        ), 400

    if not os.path.isdir(scanner_path):
        return jsonify(
            ok      = False,
            message = f'Scanner folder not found: "{scanner_path}". '
                      f'Please update the path in Settings.',
            action  = 'open_settings'
        ), 400

    # 2. Find the latest scanned file
    latest = get_latest_scanned_file(scanner_path)
    if not latest:
        return jsonify(
            ok      = False,
            message = f'No scanned files found in "{scanner_path}". '
                      f'Please scan a document first, then click this button.'
        ), 400

    # 3. Copy from scanner folder → customer folder
    try:
        doc = _save_disk_file(latest, customer_id)
        db.session.commit()

        return jsonify(
            ok          = True,
            message     = f'"{doc.original_name}" imported from scanner!',
            filename    = doc.original_name,
            folder      = sanitize_name(c.name),
            size        = doc.size_str,
            source_path = latest,
        )
    except ValueError as e:
        return jsonify(ok=False, message=str(e)), 400
    except Exception as e:
        db.session.rollback()
        return jsonify(ok=False, message=f'Server error: {str(e)}'), 500


# ─── Manual upload (browser file picker, AJAX) ────────────────────────────────
# POST /upload_scan/<customer_id>   ← kept for backward compatibility

@app.route('/upload_scan/<int:customer_id>', methods=['POST'])
@login_required
def upload_scan(customer_id):
    """Manual file-picker upload via fetch(). Returns JSON."""
    c = Customer.query.get(customer_id)
    if not c:
        return jsonify(ok=False, message='Customer not found.'), 404
    try:
        doc = _save_file_object(
            request.files.get('scan_file'),
            request.files.get('scan_file').filename if request.files.get('scan_file') else '',
            customer_id
        )
        db.session.commit()
        return jsonify(
            ok       = True,
            message  = f'"{doc.original_name}" uploaded successfully!',
            filename = doc.original_name,
            folder   = sanitize_name(c.name),
            size     = doc.size_str,
        )
    except (ValueError, AttributeError) as e:
        return jsonify(ok=False, message=str(e)), 400
    except Exception as e:
        db.session.rollback()
        return jsonify(ok=False, message=f'Server error: {str(e)}'), 500


# ─── Regular form upload ───────────────────────────────────────────────────────

@app.route('/upload/<int:customer_id>', methods=['POST'])
@login_required
def upload_document(customer_id):
    Customer.query.get_or_404(customer_id)
    f = request.files.get('document')
    try:
        _save_file_object(f, f.filename if f else '', customer_id)
        db.session.commit()
        flash('File uploaded successfully!', 'success')
    except ValueError as e:
        flash(str(e), 'error')
    return redirect(url_for('customer_profile', cid=customer_id))


# ─── Document Serve / Download / Delete ───────────────────────────────────────

@app.route('/documents/view/<int:did>')
@login_required
def view_document(did):
    doc    = Document.query.get_or_404(did)
    folder = customer_upload_dir(doc.customer_id)
    fpath  = os.path.join(folder, doc.filename)
    if not os.path.isfile(fpath):
        flash('File not found on disk.', 'error')
        return redirect(url_for('customer_profile', cid=doc.customer_id))
    mime, _ = mimetypes.guess_type(doc.filename)
    return send_from_directory(folder, doc.filename,
                               mimetype=mime or 'application/octet-stream',
                               as_attachment=False)


@app.route('/documents/download/<int:did>')
@login_required
def download_document(did):
    doc    = Document.query.get_or_404(did)
    folder = customer_upload_dir(doc.customer_id)
    fpath  = os.path.join(folder, doc.filename)
    if not os.path.isfile(fpath):
        flash('File not found on disk.', 'error')
        return redirect(url_for('customer_profile', cid=doc.customer_id))
    return send_from_directory(folder, doc.filename,
                               as_attachment=True,
                               download_name=doc.original_name)


@app.route('/documents/delete/<int:did>', methods=['POST'])
@login_required
def delete_document(did):
    doc    = Document.query.get_or_404(did)
    cid    = doc.customer_id
    folder = customer_upload_dir(cid)
    fpath  = os.path.join(folder, doc.filename)
    if os.path.isfile(fpath):
        os.remove(fpath)
    db.session.add(Activity(customer_id=cid,
                            description=f'Document deleted: {doc.original_name}'))
    db.session.delete(doc)
    db.session.commit()
    flash(f'"{doc.original_name}" deleted.', 'success')
    return redirect(url_for('customer_profile', cid=cid))


# ─── Activity ─────────────────────────────────────────────────────────────────

@app.route('/customers/<int:cid>/activity', methods=['POST'])
@login_required
def add_activity(cid):
    desc = request.form.get('description', '').strip()
    if desc:
        db.session.add(Activity(customer_id=cid, description=desc))
        db.session.commit()
        flash('Activity logged!', 'success')
    return redirect(url_for('customer_profile', cid=cid))


# ─── Billing ──────────────────────────────────────────────────────────────────

@app.route('/billing')
@login_required
def billing():
    return render_template('billing.html',
                           customers=Customer.query.order_by(Customer.name).all())


@app.route('/billing/save', methods=['POST'])
@login_required
def save_bill():
    cid       = request.form.get('customer_id')
    items_raw = request.form.get('items', '[]')
    subtotal  = float(request.form.get('subtotal', 0))
    discount  = float(request.form.get('discount', 0))
    total     = float(request.form.get('total', 0))
    inv_num   = get_next_invoice_number()
    bill      = Bill(invoice_number=inv_num, customer_id=cid, items=items_raw,
                     subtotal=subtotal, discount=discount, total=total)
    db.session.add(bill)
    db.session.add(Activity(customer_id=cid,
                            description=f'Invoice {inv_num} generated — ₹{total:.2f}'))
    db.session.commit()
    flash(f'Invoice {inv_num} saved!', 'success')
    return redirect(url_for('invoice_view', bid=bill.id))


@app.route('/invoices')
@login_required
def invoices():
    return render_template('invoices.html',
                           bills=Bill.query.order_by(Bill.date.desc()).all())


@app.route('/invoices/<int:bid>')
@login_required
def invoice_view(bid):
    bill = Bill.query.get_or_404(bid)
    return render_template('invoice_print.html',
                           bill=bill, items=json.loads(bill.items))


# ─── Excel Export ─────────────────────────────────────────────────────────────

@app.route('/export/customers')
@login_required
def export_customers():
    path = rebuild_excel()
    return send_file(path, as_attachment=True, download_name='customers.xlsx')


# ─── App Init ─────────────────────────────────────────────────────────────────

with app.app_context():
    db.create_all()
    os.makedirs(UPLOAD_BASE, exist_ok=True)
    os.makedirs(EXCEL_DIR,   exist_ok=True)

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
